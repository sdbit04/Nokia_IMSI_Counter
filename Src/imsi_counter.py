import os
import re
import gzip
import argparse
import openpyxl

# project_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# input_path = os.path.join(project_path, "Input_data\\data.bin_decoded.txt.2020-09-09-165.log.gz")


def get_imsi_vs_time_gz_no_filter(input_file_path, imsi_vs_time: dict = {}):
    imsi_exp = re.compile(rb"\s+iMSI\s+=", re.IGNORECASE)
    start_of_block = re.compile(rb"[0-9]+\s([0-9]{2,2}\s+[0-9]{2,2}\s[0-9]{2,2}\s+[0-9]{3,3})")
    input_path = input_file_path
    try:
        with gzip.open(input_path, 'r') as file_ob:
            file_lines = file_ob.readlines()
            print("Searching all the rows!!")
            for ind, line in enumerate(file_lines):
                if re.match(imsi_exp, line):
                    imsi = "{}".format(line.decode().strip()).upper().lstrip("IMSI = ").lstrip(" ")
                    # TODO if imsi into the filter
                    imsi_time = None
                    for i in range(ind - 7, ind - 600, -1):
                        start_block = re.match(start_of_block, file_lines[i])
                        if start_block is not None:
                            imsi_time = start_block.group(1)
                            if imsi_time is not None:
                                imsi_time = imsi_time.decode().replace(r"\t", "-")

                            break
                    if imsi not in imsi_vs_time.keys():
                        imsi_vs_time[imsi] = [imsi_time]
                    else:
                        print("Multiple entries for IMSI ={}".format(imsi))
                        imsi_vs_time[imsi].append(imsi_time)
                else:
                    continue
    except:
        print("Got exception while reading {}".format(input_path))
        return imsi_vs_time
    return imsi_vs_time


def get_imsi_vs_time_gz_one_imsi(input_file_path, imsi, imsi_vs_time: dict = {}):
    # imsi_exp = re.compile(rb"\s+iMSI\s+=", re.IGNORECASE)
    _imsi = imsi
    imsi_reg = "\s+iMSI\s+=\s+{0}".format(imsi)
    imsi_exp = re.compile(imsi_reg.encode(), re.IGNORECASE)
    start_of_block = re.compile(rb"[0-9]+\s([0-9]{2,2}\s+[0-9]{2,2}\s[0-9]{2,2}\s+[0-9]{3,3})")
    input_path = input_file_path
    try:
        with gzip.open(input_path, 'r') as file_ob:
            file_lines = file_ob.readlines()
            print("Searching all the rows!!")
            for ind, line in enumerate(file_lines):
                if re.match(imsi_exp, line):
                    # imsi = "{}".format(line.decode().strip()).upper().lstrip("IMSI = ").lstrip(" ")
                    # TODO if imsi into the filter
                    imsi_time = None
                    for i in range(ind - 7, ind - 600, -1):
                        start_block = re.match(start_of_block, file_lines[i])
                        if start_block is not None:
                            imsi_time = start_block.group(1)
                            if imsi_time is not None:
                                imsi_time = imsi_time.decode().replace(r"\t", "-")

                            break
                    if imsi not in imsi_vs_time.keys():
                        imsi_vs_time[imsi] = [imsi_time]
                    else:
                        print("Multiple entries for IMSI ={}".format(imsi))
                        imsi_vs_time[imsi].append(imsi_time)
                else:
                    continue
    except:
        print("Got exception while reading {}".format(input_path))
        return imsi_vs_time
    return imsi_vs_time


def get_imsi_vs_time_gz(input_file_path, imsi_list, imsi_vs_time: dict = {}):
    imsi_exp = re.compile(rb"\s+iMSI\s+=", re.IGNORECASE)
    start_of_block = re.compile(rb"[0-9]+\s([0-9]{2,2}\s+[0-9]{2,2}\s[0-9]{2,2}\s+[0-9]{3,3})")
    input_path = input_file_path
    imsi_list = imsi_list
    try:
        with gzip.open(input_path, 'r') as file_ob:
            file_lines = file_ob.readlines()
            print("Searching all the rows!!")
            for ind, line in enumerate(file_lines):
                if re.match(imsi_exp, line):
                    imsi = "{}".format(line.decode().strip()).upper().lstrip("IMSI = ").lstrip(" ")
                    # TODO if imsi into the filter
                    # print("imsi = {}".format(imsi))
                    if imsi in imsi_list:
                        print("IMSI matched! with IMSI filter")
                        imsi_time = None
                        for i in range(ind-7, ind-600, -1):
                            start_block = re.match(start_of_block, file_lines[i])
                            if start_block is not None:
                                imsi_time = start_block.group(1)
                                if imsi_time is not None:
                                    imsi_time = imsi_time.decode().replace(r"\t", "-")
                                break
                        if imsi not in imsi_vs_time.keys():
                            imsi_vs_time[imsi] = [imsi_time]
                        else:
                            print("Multiple entries for IMSI ={}".format(imsi))
                            imsi_vs_time[imsi].append(imsi_time)
                    else:
                        continue

    except:
        print("Got exception while reading {}".format(input_path))
        return imsi_vs_time
    return imsi_vs_time


def get_imsi_vs_time(input_dir, imsi_list, total_imsi_v_time={}):
    _input_dir = input_dir
    _imsi_list = imsi_list
    if _imsi_list is not None:
        if len(_imsi_list) == 1:
            imsi = list(imsi_list)[0].upper()
            for p, d, files in os.walk(_input_dir):
                for file in files:
                    f_path = os.path.join(p, file)
                    print(f_path)
                    total_imsi_v_time = get_imsi_vs_time_gz_one_imsi(f_path, imsi, total_imsi_v_time)
        else:
            for p, d, files in os.walk(_input_dir):
                for file in files:
                    f_path = os.path.join(p, file)
                    print(f_path)
                    total_imsi_v_time = get_imsi_vs_time_gz(f_path, _imsi_list, total_imsi_v_time)
    else:
        for p, d, files in os.walk(_input_dir):
            for file in files:
                f_path = os.path.join(p, file)
                print(f_path)
                total_imsi_v_time = get_imsi_vs_time_gz_no_filter(f_path, total_imsi_v_time)

    return total_imsi_v_time


def read_config_file(file_path):
    config_dict = {}
    with open(file_path, 'r') as config_ob:
        lines = config_ob.readlines()
        for line in lines:
            input_dir = re.match(re.compile("input_dir", re.IGNORECASE), line)
            if input_dir is not None:

                input_dir = line.strip().split("=")[1].strip()
                config_dict["input_dir"]=input_dir
            output_dir = re.match(re.compile("output_dir", re.IGNORECASE), line)
            if output_dir is not None:
                output_dir = line.strip().split("=")[1].strip()
                config_dict["output_dir"]=output_dir
            imsi_list = re.match(re.compile("imsi_list", re.IGNORECASE), line)
            if imsi_list is not None:
                imsi_list = line.strip().split("=")[1].strip().upper()
                config_dict["imsi_list"]=set(imsi_list.split(","))
    return config_dict


def write_to_out(out_dir, imsi_v_time):
    out_dir = out_dir
    imsi_v_time = imsi_v_time
    wb = openpyxl.Workbook()
    sheet = wb.worksheets[0]
    sheet.cell(1, 1, value="IMSI")
    sheet.cell(1, 2, value="Time (H24-Min-Sec-Miliseconds)")
    f_row = 1
    for key in imsi_v_time.keys():
        values = imsi_v_time[key]
        for value in values:
            if value is not None:
                f_row += 1
                # print("row = {}, value = {}".format(f_row, value))
                sheet.cell(f_row, 1, value=key)
                sheet.cell(f_row, 2, value=value.replace("\t", "-"))
    wb.save(os.path.join(out_dir, "output.xlsx"))


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("config_file_path", help="Please provide config file path")
    arguments = parser.parse_args()
    config_path = arguments.config_file_path
    config_dict = read_config_file(config_path)
    input_dir = config_dict["input_dir"]
    output_directory = config_dict["output_dir"]
    imsi_list = config_dict.get("imsi_list")
    print(config_dict)
    print("imsi_list = {}".format(imsi_list))
    try:
        total_imsi_v_time = get_imsi_vs_time(input_dir, imsi_list)
        write_to_out(output_directory, total_imsi_v_time)
    except:
        print("Exception occurs, need to handle")
    else:
        print("Success!")

